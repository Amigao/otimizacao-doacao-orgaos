import openpyxl

def ler_matriz_adjacencias(arquivo, planilha_nome):
    # Abre o arquivo Excel
    wb = openpyxl.load_workbook(arquivo)
    ws = wb[planilha_nome]
    
    # Descobre o tamanho n da matriz (quantidade de colunas a partir da coluna B)
    n = 0
    col = 2  # Começando na coluna 'B'
    while ws.cell(row=1, column=col).value is not None:
        n += 1
        col += 1

    # Inicializa a matriz de adjacências
    matriz = [[0 for _ in range(n)] for _ in range(n)]

    # Lê os valores da planilha e preenche a matriz
    for i in range(n):
        for j in range(n):
            valor = ws.cell(row=i+2, column=j+2).value
            matriz[i][j] = valor

    return matriz
